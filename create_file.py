from openpyxl import Workbook

# Create a new workbook
wb = Workbook()
ws = wb.active

# Add some data
ws['A1'] = "Name"
ws['B1'] = "Age"
ws.append(["Alice", 25])
ws.append(["Bob", 30])

# Save as .xlsx
wb.save("data.xlsx")